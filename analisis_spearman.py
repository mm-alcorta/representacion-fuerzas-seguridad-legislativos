#---------------------------------
# Importación de Librerías
#---------------------------------

import pandas as pd
from scipy.stats import spearmanr
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

#-----------------------------------------
# Definición de información del proyecto
#-----------------------------------------

archivo = "./base_datos_brasil_2022_argentina_2023.xlsx"

conjuntos_datos = [

# ---------------- BRASIL ----------------

    {
        "hoja": "Brasil",
        "nombre": "Policía Civil (2022)",
        "pais": "Brasil",
        "variable_dependiente": "Número de legisladores electos con pasado en la Policía Civil (2022)"
    },

    {
        "hoja": "Brasil",
        "nombre": "Policía Militar (2022)",
        "pais": "Brasil",
        "variable_dependiente": "Número de legisladores electos con antecedentes en la Policía Militar (2022)"
    },

    {
        "hoja": "Brasil",
        "nombre": "Fuerzas Armadas en retiro (2022)",
        "pais": "Brasil",
        "variable_dependiente": "Número de legisladores electos con procedencia institucional en las Fuerzas Armadas en retiro (2022)"
    },

    {
        "hoja": "Brasil",
        "nombre": "Fuerzas Armadas en servicio activo (2022)",
        "pais": "Brasil",
        "variable_dependiente": "Número de legisladores electos con procedencia institucional en las Fuerzas Armadas en servicio activo (2022)"
    },

# ---------------- ARGENTINA ----------------

    {
        "hoja": "Argentina",
        "nombre": "Policía Provincial (2023)",
        "pais": "Argentina",
        "variable_dependiente": "Número de legisladores electos con experiencia previa en la Policía Provincial (2023)"
    },

    {
        "hoja": "Argentina",
        "nombre": "Fuerzas Armadas (2023)",
        "pais": "Argentina",
        "variable_dependiente": "Número de legisladores electos con procedencia institucional en las Fuerzas Armadas (2023)"
    }

]

# ----------------------------------
# Carga y preparación de los datos
# ----------------------------------

hojas = {
    "Brasil": pd.read_excel(archivo, sheet_name="Brasil"),
    "Argentina": pd.read_excel(archivo, sheet_name="Argentina")
}

for df in hojas.values():
    df.columns = df.columns.str.strip()

# -----------------------------------------------------
# Funciones para la interpretación de los resultados
# -----------------------------------------------------

def obtener_intensidad_asociacion(rho):
    valor_absoluto = abs(rho)

    if valor_absoluto < 0.20:
        return "Muy débil"
    elif valor_absoluto < 0.40:
        return "Débil"
    elif valor_absoluto < 0.60:
        return "Moderada"
    elif valor_absoluto < 0.80:
        return "Fuerte"
    else:
        return "Muy fuerte"


def obtener_significancia_estadistica(valor_p):
    if valor_p < 0.05:
        return "Significativa"
    else:
        return "No significativa"

# ----------------------------------------
# Cálculo de la correlación de Spearman
# ----------------------------------------

print("Iniciando análisis de correlación de Spearman...")

resultados = []

for conjunto in conjuntos_datos:

    df = hojas[conjunto["hoja"]]

    # Selección de indicadores según el país analizado

    if conjunto["pais"] == "Brasil":
        indicadores = [
            "Índice de Gini per cápita de los hogares (2021)",
            "Tasa de homicidios por cada 100.000 habitantes (2021)"
        ]
    else:
        indicadores = [
            "Índice de Gini del ingreso per cápita familiar (tercer trimestre 2022)",
            "Tasa de homicidios por cada 100.000 habitantes (2022)"
        ]

    for indicador in indicadores:

        # Eliminación de valores faltantes

        datos_validos = df[
            [
                conjunto["variable_dependiente"],
                indicador
            ]
        ].dropna()

        # Estimación de la correlación y significancia estadística

        rho, valor_p = spearmanr(
            datos_validos[conjunto["variable_dependiente"]],
            datos_validos[indicador]
        )

        # Almacenamiento de resultados

        resultados.append({

            "País": conjunto["pais"],

            "Legisladores electos provenientes de las fuerzas públicas de seguridad": conjunto["nombre"],

            "Indicadores": indicador,

            "Tamaño de la muestra": len(datos_validos),

            "Coeficiente rho": rho,

            "Valor p": valor_p,

            "Intensidad de la asociación": obtener_intensidad_asociacion(rho),

            "Significancia estadística": obtener_significancia_estadistica(valor_p)

        })

#--------------------------------------------
# Construcción del Dataframe de resultados
#--------------------------------------------

df_resultados = (
    pd.DataFrame(resultados)
    .sort_values(by=["País"])
    .reset_index(drop=True)
)

# ------------------------------------
# Exportación de resultados a Excel
# ------------------------------------

df_resultados.to_excel("resultados_spearman.xlsx", index=False)

wb = load_workbook("resultados_spearman.xlsx")
ws = wb.active

# Ajuste automático del ancho de las columnas

for columna in ws.columns:
    max_len = 0
    letra = columna[0].column_letter

    for celda in columna:
        if celda.value:
            max_len = max(max_len, len(str(celda.value)))

    ws.column_dimensions[letra].width = max_len + 2

# Aplicación de bordes a todas las celdas

borde = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for fila in ws.iter_rows():
    for celda in fila:
        celda.border = borde

wb.save("resultados_spearman.xlsx")

#----------------------------
# Finalización del proceso
#----------------------------

print("\nAnálisis finalizado correctamente.")
print("Resultados exportados a 'resultados_spearman.xlsx'.")